import openpyxl
from openpyxl.styles import Font
from copy import copy
from datetime import date
import pandas as pd

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def get_mapping(headers):
    """
    Dynamically maps dataframe column names to template headers.
    """
    standard_map = {
        "ServiceNow Ticket #": "Ticket No.",
        "Ticket No.": "Ticket No.",
        "REQ No.": "REQ No.",
        "Type": "Type",
        "Description": "Request Detail",
        "Request Detail": "Request Detail",
        "Requested for": "Requested for",
        "PIC": "Assign To",
        "Assign To": "Assign To",
        "Received": "Time - Arrive",
        "Time - Arrive": "Time - Arrive",
        "Resolved": "Time - Close",
        "Time - Close": "Time - Close",
        "Remarks": "Remarks"
    }
    return {standard_map.get(h, h): col_idx for col_idx, h in headers.items()}

def find_header_and_cols(ws, anchor_text):
    """
    Finds the row and column mapping for a given section start header.
    Only captures contiguous headers to avoid section collisions.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == anchor_text:
                header_row = cell.row
                anchor_col = cell.column
                
                headers = {}
                # Look right from anchor (inclusive)
                for col in range(anchor_col, 27):
                    val = ws.cell(row=header_row, column=col).value
                    if not val: break
                    headers[col] = str(val).strip()
                
                # Look left from anchor
                for col in range(anchor_col - 1, 0, -1):
                    val = ws.cell(row=header_row, column=col).value
                    if not val: break
                    headers[col] = str(val).strip()
                
                return header_row, get_mapping(headers)
    return None, None

def write_v2_report(
    template_path: str,
    output_path: str,
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    summary: dict,
    begin_date: date,
    end_date: date
):
    wb = openpyxl.load_workbook(template_path)
    ws = wb["IFS"] if "IFS" in wb.sheetnames else wb.active

    # 0. Header texts
    title_font = Font(size=22)
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=5)
    ws.cell(row=3, column=2).value = "IFS AMS Workload Summary"
    ws.cell(row=3, column=2).font = title_font
    ws.cell(row=5, column=2).value = "For NAFTA Marelli USA"
    ws.cell(row=8, column=3).value = "Open Items"

    # 1. Update Period / Summary
    period_label_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Period:":
                period_label_cell = cell
                break
        if period_label_cell: break
    
    if period_label_cell:
        # Relocate Period label/value to row 7
        label_col = period_label_cell.column
        ws.cell(row=period_label_cell.row, column=label_col).value = None
        ws.cell(row=period_label_cell.row, column=label_col + 1).value = None

        ws.cell(row=7, column=label_col).value = "Period:"
        target_cell = ws.cell(row=7, column=label_col + 1)
        target_cell.value = f"{begin_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')} {summary['closed_count']} closed, {summary['open_count']} open."

    # 2. Helper to write section
    def write_section(df, anchor_header, allowed_cols=None):
        header_row, col_map = find_header_and_cols(ws, anchor_header)
        if not header_row:
            print(f"Warning: Header '{anchor_header}' not found in template.")
            return

        # Filter col_map to only allowed columns if specified
        if allowed_cols:
            col_map = {k: v for k, v in col_map.items() if k in allowed_cols}

        start_row = header_row + 1
        # Style source: first data row (for newly extended rows only)
        style_source = {col_idx: ws.cell(row=start_row, column=col_idx) for col_idx in col_map.values()}

        # Clear existing values in the section range (preserve formatting)
        existing_max_row = ws.max_row
        for row_idx in range(start_row, existing_max_row + 1):
            for sheet_col in col_map.values():
                ws.cell(row=row_idx, column=sheet_col).value = None

        for i, (_, row) in enumerate(df.iterrows()):
            curr_row = start_row + i
            for source_col, sheet_col in col_map.items():
                cell = ws.cell(row=curr_row, column=sheet_col)
                # If writing past existing rows, copy style from template row
                if curr_row > existing_max_row:
                    copy_cell_style(style_source[sheet_col], cell)
                # Just write value - preserve all template formatting
                val = row.get(source_col)
                
                # For Time - Arrive and Time - Close columns, extract only the date
                if source_col in ["Time - Arrive", "Time - Close"]:
                    if pd.notna(val):
                        # Convert to date only (remove time)
                        if isinstance(val, pd.Timestamp):
                            val = val.date()
                
                cell.value = val

    # Write sections
    # Left section: Open tickets in date period (Status = "OPEN")
    left_status = left_df["Status"].astype(str).str.strip().str.upper()
    left_df_open = left_df[left_status == "OPEN"].copy()
    # Only keep Ticket No. and Request Detail columns for left section
    write_section(left_df_open, "Ticket No.", allowed_cols=["Ticket No.", "Request Detail"])
    
    # Right section: Open backlog (current/previous year, Status = "OPEN")
    write_section(right_df, "ServiceNow Ticket #")

    wb.save(output_path)
