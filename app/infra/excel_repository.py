import pandas as pd
import openpyxl
from openpyxl.styles import Font
from copy import copy
from datetime import date
from typing import Dict, Any, List

def load_excel_data(file_path: str) -> Dict[str, pd.DataFrame]:
    """Pure I/O: Loads all sheets from an Excel file."""
    return pd.read_excel(file_path, sheet_name=None)

def copy_cell_style(source_cell, target_cell):
    """Utility for Excel writing."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def write_report_to_template(
    template_path: str,
    output_path: str,
    sections: List[Dict[str, Any]],
    summary_info: Dict[str, Any]
):
    """
    Pure I/O: Writes prepared data to an Excel template.
    Expects data already prepared by the service layer.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb["IFS"] if "IFS" in wb.sheetnames else wb.active

    # 1. Header/Summary Writing
    ws.cell(row=3, column=2).value = "IFS AMS Workload Summary"
    ws.cell(row=3, column=2).font = Font(size=22)
    ws.cell(row=5, column=2).value = "For NAFTA Marelli USA"
    
    # Locate period/summary cell and update
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Period:":
                label_col = cell.column
                # Standard V2 template locations
                ws.cell(row=7, column=label_col).value = "Period:"
                ws.cell(row=7, column=label_col + 1).value = f"{summary_info['period']} {summary_info['stats']}"
                break

    # 2. Section Writing
    # Sections are passed as a list of dicts: {anchor: str, df: DataFrame, allowed_cols: list}
    for section in sections:
        anchor = section['anchor']
        df = section['df']
        allowed_cols = section.get('allowed_cols')
        
        # Implementation of finding and writing to the section (moved from writer service)
        header_row, col_map = _find_header_mapping(ws, anchor, allowed_cols)
        if header_row:
             _write_df_to_ws(ws, df, header_row + 1, col_map)

    wb.save(output_path)

def _find_header_mapping(ws, anchor_text: str, allowed_cols: List[str] = None) -> (int, dict):
    """Finds header row and column map for a specific anchor."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == anchor_text:
                header_row = cell.row
                mapping = {}
                # Capture headers to the right/left (simplified version of writing logic)
                for col in range(cell.column - 5, cell.column + 20):
                    if col <= 0: continue
                    val = ws.cell(row=header_row, column=col).value
                    if val:
                        col_name = str(val).strip()
                        if not allowed_cols or col_name in allowed_cols:
                            mapping[col_name] = col
                return header_row, mapping
    return None, None

def _write_df_to_ws(ws, df, start_row, col_map):
    """Internal helper to write a dataframe to a worksheet starting at a row."""
    existing_max_row = ws.max_row
    # First clear existing data range below headers
    for row_idx in range(start_row, existing_max_row + 200): # Buffer clear
        for col_idx in col_map.values():
            ws.cell(row=row_idx, column=col_idx).value = None

    # Style source: first data row if it exists
    style_source = {col_idx: ws.cell(row=start_row, column=col_idx) for col_idx in col_map.values()}

    for i, (_, row) in enumerate(df.iterrows()):
        curr_row = start_row + i
        for col_name, col_idx in col_map.items():
            cell = ws.cell(row=curr_row, column=col_idx)
            if curr_row > existing_max_row:
                copy_cell_style(style_source[col_idx], cell)
            
            val = row.get(col_name)
            if isinstance(val, (pd.Timestamp, date)):
                val = val.date() if hasattr(val, 'date') else val
            cell.value = val
