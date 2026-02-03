"""
Weekly Report Generator - Core Logic Module
============================================

Refactored from generate_report.py to work with Streamlit
Removes command-line prompts and returns file buffers
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import io
from copy import copy


class WeeklyReportGenerator:
    """Generate weekly ServiceNow reports from Excel data"""
    
    # Configuration constants
    SOURCE_SHEET_TICKETS = "2025"
    SOURCE_SHEET_USERS = "New Users"
    TARGET_SHEET_TICKETS = "IFS"
    TARGET_SHEET_USERS = "New User"
    DATA_START_ROW = 3
    
    def get_last_week_date_range(self, reference_date=None):
        """Calculate the date range for the previous week"""
        if reference_date is None:
            reference_date = datetime.today()
        start_of_last_week = reference_date - timedelta(days=reference_date.weekday() + 7)
        start_of_last_week = start_of_last_week.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_last_week = start_of_last_week + timedelta(days=6, hours=23, minutes=59, seconds=59)
        return start_of_last_week, end_of_last_week
    
    def clean_and_get_styles(self, wb, sheet_name, start_row):
        """Clean sheet data and extract cell styles"""
        styles = {}
        if sheet_name not in wb.sheetnames:
            return styles
        
        ws = wb[sheet_name]
        
        # Extract styles from header row
        for col in range(1, 27):
            cell = ws.cell(row=start_row, column=col)
            styles[col] = {
                'font': copy(cell.font),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'number_format': copy(cell.number_format),
                'alignment': copy(cell.alignment)
            }
        
        # Clean data
        max_row = ws.max_row
        if sheet_name == self.TARGET_SHEET_TICKETS:
            # Clean data table (H-Z)
            if max_row >= start_row:
                for row in range(start_row, max_row + 1):
                    for col in range(8, 27):
                        try:
                            ws.cell(row=row, column=col).value = None
                        except:
                            pass
            
            # Clean dashboard list (A-G) from Row 9
            DASHBOARD_START_ROW = 9
            if max_row >= DASHBOARD_START_ROW:
                for row in range(DASHBOARD_START_ROW, max_row + 1):
                    for col in range(1, 8):
                        try:
                            ws.cell(row=row, column=col).value = None
                        except:
                            pass
        else:
            # Standard cleaning for other sheets
            if max_row >= start_row:
                for row in range(start_row, max_row + 1):
                    for col in range(1, 27):
                        try:
                            ws.cell(row=row, column=col).value = None
                        except:
                            pass
        
        return styles
    
    def apply_style(self, cell, style_dict):
        """Apply saved styles to a cell"""
        if not style_dict:
            return
        try:
            cell.font = copy(style_dict['font'])
            cell.border = copy(style_dict['border'])
            cell.fill = copy(style_dict['fill'])
            cell.number_format = copy(style_dict['number_format'])
            cell.alignment = copy(style_dict['alignment'])
        except:
            pass
    
    def get_col(self, df, candidates):
        """Find first matching column from candidates"""
        for c in candidates:
            if c in df.columns:
                return c
        return None
    
    def generate_report(self, source_file, template_file, reference_date=None, source_sheet="2025"):
        """
        Generate weekly report
        
        Args:
            source_file: Uploaded source Excel file (file-like object)
            template_file: Uploaded template Excel file (file-like object)
            reference_date: Reference date for week calculation
            source_sheet: Name of the source sheet to read
        
        Returns:
            BytesIO buffer containing the generated Excel file
        """
        if reference_date is None:
            reference_date = datetime.today()
        
        # Calculate date range
        start_date, end_date = self.get_last_week_date_range(reference_date)
        
        # Load template into memory
        template_buffer = io.BytesIO(template_file.read())
        wb = openpyxl.load_workbook(template_buffer)
        
        # Clean and get styles
        ifs_styles = self.clean_and_get_styles(wb, self.TARGET_SHEET_TICKETS, self.DATA_START_ROW)
        user_styles = self.clean_and_get_styles(wb, self.TARGET_SHEET_USERS, self.DATA_START_ROW)
        
        # Update dashboard headers
        ws_ifs = wb[self.TARGET_SHEET_TICKETS]
        ws_ifs['C3'] = "IFS AMS Workload Summary"
        ws_ifs['D5'] = "For NAFTA Marelli USA"
        
        # Process tickets
        sheets_to_load = list(dict.fromkeys([source_sheet, "2025", "2026"]))
        df_list = []
        
        for sheet in sheets_to_load:
            try:
                df_temp = pd.read_excel(source_file, sheet_name=sheet)
                df_temp.columns = df_temp.columns.astype(str).str.strip()
                df_list.append(df_temp)
            except:
                pass
        
        if not df_list:
            raise ValueError("No data sheets found in source file")
        
        # Combine and remove duplicates
        df_tickets = pd.concat(df_list, ignore_index=True)
        df_tickets = df_tickets.drop_duplicates()
        
        # Normalize date column
        renamed = False
        for c in df_tickets.columns:
            if c.lower() in ['date', 'created', 'opened', 'start date']:
                df_tickets.rename(columns={c: 'Date Created'}, inplace=True)
                renamed = True
                break
        
        if not renamed:
            for c in df_tickets.columns:
                if 'date' in c.lower():
                    df_tickets.rename(columns={c: 'Date Created'}, inplace=True)
                    renamed = True
                    break
        
        if 'Date Created' not in df_tickets.columns and len(df_tickets.columns) > 1:
            df_tickets.rename(columns={df_tickets.columns[1]: 'Date Created'}, inplace=True)
        
        if 'Date Created' not in df_tickets.columns:
            raise ValueError("Date column not found in source data")
        
        df_tickets['Date Created'] = pd.to_datetime(df_tickets['Date Created'], errors='coerce')
        
        # Find status column
        status_col = None
        for c in df_tickets.columns:
            if 'status' in c.lower():
                status_col = c
                break
        
        # Filter logic
        mask_date_valid = df_tickets['Date Created'] <= end_date
        
        if status_col:
            status_clean = df_tickets[status_col].astype(str).str.strip().str.lower()
            mask_open = (status_clean == "open")
        else:
            mask_open = pd.Series([False] * len(df_tickets))
        
        mask_week = (df_tickets['Date Created'] >= start_date) & (df_tickets['Date Created'] <= end_date)
        
        filtered_tickets = df_tickets[mask_open & mask_date_valid]
        
        # Calculate stats
        df_created_in_week = df_tickets[mask_week]
        if status_col:
            stats_status = df_created_in_week[status_col].astype(str).str.strip().str.lower()
            count_closed = stats_status[stats_status.str.contains('close')].count()
            count_open = stats_status[stats_status == 'open'].count()
        else:
            count_closed = 0
            count_open = 0
        
        period_str = f"Period: {start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}   {count_closed} closed, {count_open} open"
        
        # Update dashboard
        ws_ifs['D7'].value = None
        ws_ifs['E7'].value = None
        ws_ifs['D7'] = period_str
        
        # Build column mapping
        ws_tickets = wb[self.TARGET_SHEET_TICKETS]
        template_headers = {}
        for col_idx in range(1, 27):
            header_cell = ws_tickets.cell(row=2, column=col_idx)
            if header_cell.value:
                header_name = str(header_cell.value).strip()
                template_headers[col_idx] = header_name
        
        map_tickets = {}
        
        # Map columns
        for template_col_idx, template_header in template_headers.items():
            template_header_clean = template_header.strip()
            
            if template_header_clean in ['ServiceNow Ticket #', 'ServiceNow Ticket']:
                source_col = self.get_col(df_tickets, ['Ticket No.', 'Ticket No', 'Ticket'])
                if source_col:
                    map_tickets[template_col_idx] = source_col
            elif template_header_clean == 'Description':
                if 'Request Detail' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Request Detail'
            elif template_header_clean == 'Type':
                if 'Type' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Type'
            elif template_header_clean == 'PIC':
                if 'Assign To' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Assign To'
            elif template_header_clean == 'Received':
                if 'Time - Arrive' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Time - Arrive'
            elif template_header_clean == 'Resolved':
                if 'Time - Close' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Time - Close'
            elif template_header_clean in df_tickets.columns:
                map_tickets[template_col_idx] = template_header_clean
        
        # Write ticket data
        row_idx = self.DATA_START_ROW
        for _, row in filtered_tickets.iterrows():
            for col_idx in range(8, 27):
                cell = ws_tickets.cell(row=row_idx, column=col_idx)
                self.apply_style(cell, ifs_styles.get(col_idx))
                if col_idx in map_tickets and map_tickets[col_idx]:
                    val = row.get(map_tickets[col_idx])
                    if val is not None:
                        cell.value = val
            row_idx += 1
        
        row_idx_tickets = row_idx
        
        # Process users
        df_users = pd.read_excel(source_file, sheet_name=self.SOURCE_SHEET_USERS)
        df_users.columns = df_users.columns.astype(str).str.strip()
        
        renamed_u = False
        for c in df_users.columns:
            if 'date' in c.lower():
                df_users.rename(columns={c: 'Date Created'}, inplace=True)
                renamed_u = True
                break
        
        if not renamed_u and len(df_users.columns) > 1:
            df_users.rename(columns={df_users.columns[1]: 'Date Created'}, inplace=True)
        
        df_users['Date Created'] = pd.to_datetime(df_users['Date Created'], errors='coerce')
        
        mask_date_u = (df_users['Date Created'] >= start_date) & (df_users['Date Created'] <= end_date)
        
        cat_col = None
        for c in df_users.columns:
            if 'category' in c.lower():
                cat_col = c
                break
        
        if cat_col:
            mask_cat = df_users[cat_col].astype(str).str.strip().str.lower() == "create account"
        else:
            try:
                mask_cat = df_users.iloc[:, 14].astype(str).str.strip().str.lower() == "create account"
            except:
                mask_cat = pd.Series([True] * len(df_users))
        
        filtered_users = df_users[mask_date_u & mask_cat]
        
        # Write user data
        ws_users = wb[self.TARGET_SHEET_USERS]
        row_idx = self.DATA_START_ROW
        for _, row in filtered_users.iterrows():
            for col_idx in range(1, 20):
                cell = ws_users.cell(row=row_idx, column=col_idx)
                self.apply_style(cell, user_styles.get(col_idx))
            for i, val in enumerate(row):
                cell = ws_users.cell(row=row_idx, column=i+1)
                cell.value = val
            row_idx += 1
        
        # Add auto-filters
        last_row_map = {
            self.TARGET_SHEET_TICKETS: row_idx_tickets - 1,
            self.TARGET_SHEET_USERS: row_idx
        }
        
        for ws_name, last_data_row in last_row_map.items():
            if ws_name not in wb.sheetnames:
                continue
            ws = wb[ws_name]
            
            # Remove tables
            if hasattr(ws, 'tables') and ws.tables:
                for tbl_name in list(ws.tables.keys()):
                    del ws.tables[tbl_name]
            
            # Add auto-filter
            fill_max_row = max(last_data_row, 2)
            ws.auto_filter.ref = f"A2:P{fill_max_row}"
        
        # Save to buffer
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer
