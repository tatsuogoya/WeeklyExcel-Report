import openpyxl
from openpyxl.styles import Border, Side

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IFS"
    
    # Dashboard area
    ws["B4"] = "Period:"
    # ws["C4"] will be the value
    
    # --- Left Section Headers (A-F) ---
    left_headers = [
        "Ticket No.", "REQ No.", "Type", "Request Detail", "Requested for", "Assign To"
    ]
    for col, header in enumerate(left_headers, start=1):
        ws.cell(row=10, column=col).value = header
    
    # --- Right Section Headers (H-P) ---
    right_headers = [
        "ServiceNow Ticket #", "REQ No.", "Type", "Description", 
        "Requested for", "PIC", "Received", "Resolved", "Remarks"
    ]
    for col, header in enumerate(right_headers, start=8): # H is 8
        ws.cell(row=10, column=col).value = header
    
    # Add dummy data rows for style
    thin = Side(border_style="thin", color="000000")
    # Left styles in row 11
    for col in range(1, 7):
        ws.cell(row=11, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # Right styles in row 11
    for col in range(8, 17):
        ws.cell(row=11, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Note: User moved template to root
    wb.save("weekly_template.xlsx")
    print("Template created at weekly_template.xlsx")

if __name__ == "__main__":
    create_template()
