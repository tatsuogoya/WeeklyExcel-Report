import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import shutil
from datetime import datetime, timedelta
import os
from copy import copy
import sys

# --- Configuration ---
SOURCE_FILE = "NA Daily work.xlsx"
TEMPLATE_FILE = "SNOW_report_Template.xlsx"
SOURCE_SHEET_TICKETS = "2025"
SOURCE_SHEET_USERS = "New Users"
TARGET_SHEET_TICKETS = "IFS"
TARGET_SHEET_USERS = "New User"

def get_last_week_date_range(reference_date=None):
    if reference_date is None:
        reference_date = datetime.today()
    start_of_last_week = reference_date - timedelta(days=reference_date.weekday() + 7)
    start_of_last_week = start_of_last_week.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_last_week = start_of_last_week + timedelta(days=6, hours=23, minutes=59, seconds=59)
    return start_of_last_week, end_of_last_week

def clean_and_get_styles(wb, sheet_name, start_row):
    styles = {}
    if sheet_name not in wb.sheetnames: return styles
    ws = wb[sheet_name]
    for col in range(1, 27):
        cell = ws.cell(row=start_row, column=col)
        styles[col] = {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': copy(cell.number_format),
            'alignment': copy(cell.alignment)
        }
    # NEW CLEANING STRATEGY: Split Ranges
    max_row = ws.max_row
    if sheet_name == TARGET_SHEET_TICKETS:
        # 1. Clean Data Table (H-Z) from start_row (3)
        if max_row >= start_row:
            print(f"Cleaning '{sheet_name}': Data Table (Cols 8-26)...")
            for row in range(start_row, max_row + 1):
                for col in range(8, 27):
                    try: ws.cell(row=row, column=col).value = None
                    except: pass
        
        # 2. Clean Dashboard List (A-G) from Row 9 (Preserving Headers in 1-8)
        #    Rows 3, 5, 7 have our headers. Row 9 seems to be where list starts.
        DASHBOARD_START_ROW = 9
        if max_row >= DASHBOARD_START_ROW:
            print(f"Cleaning '{sheet_name}': Dashboard List (Cols 1-7) from Row {DASHBOARD_START_ROW}...")
            for row in range(DASHBOARD_START_ROW, max_row + 1):
                for col in range(1, 8):
                    try: ws.cell(row=row, column=col).value = None
                    except: pass
                    
    else:
        # Standard Cleaning for other sheets (New User)
        clean_col_start = 1
        if max_row >= start_row:
             print(f"Cleaning '{sheet_name}': Full Sheet {start_row}-{max_row}...")
             for row in range(start_row, max_row + 1):
                 for col in range(clean_col_start, 27):
                     try: ws.cell(row=row, column=col).value = None
                     except: pass
    
    # Remove lingering tables/filters handled in main loop
    return styles

def apply_style(cell, style_dict):
    if not style_dict: return
    try:
        cell.font = copy(style_dict['font'])
        cell.border = copy(style_dict['border'])
        cell.fill = copy(style_dict['fill'])
        cell.number_format = copy(style_dict['number_format'])
        cell.alignment = copy(style_dict['alignment'])
    except: pass

def generate_report():
    print("--- Starting Weekly Report Generation ---")
    
    target_year = input(f"Enter Source Tab Name (default '{SOURCE_SHEET_TICKETS}'): ").strip()
    if not target_year: target_year = SOURCE_SHEET_TICKETS
    
    date_input = input("Enter Reference Date (YYYY-MM-DD, default Today): ").strip()
    if date_input:
        try: ref_date = datetime.strptime(date_input, "%Y-%m-%d")
        except ValueError: print("Invalid date. Using Today."); ref_date = datetime.today()
    else: ref_date = datetime.today()
    
    print(f"DEBUG: Using Reference Date: {ref_date}")

    if not os.path.exists(SOURCE_FILE): print("Source not found."); return
    if not os.path.exists(TEMPLATE_FILE): print("Template not found."); return

    start_date, end_date = get_last_week_date_range(ref_date)
    print(f"Processing Data for Week: {start_date} to {end_date}")
    
    output_filename = f"Weekly_Report_{end_date.strftime('%Y-%m-%d')}.xlsx"
    
    while True:
        try:
            shutil.copyfile(TEMPLATE_FILE, output_filename)
            break
        except PermissionError:
            print(f"ERROR: '{output_filename}' is open. Please CLOSE it and press Enter.")
            input()
        except Exception as e: print(e); return
    
    try:
        wb = openpyxl.load_workbook(output_filename)
        DATA_START_ROW = 3
        ifs_styles = clean_and_get_styles(wb, TARGET_SHEET_TICKETS, DATA_START_ROW)
        user_styles = clean_and_get_styles(wb, TARGET_SHEET_USERS, DATA_START_ROW)
        
        # --- DASHBOARD UPDATE ---
        # Since we are NOT clearing A-G anymore, we just UPDATE the values.
        # Styles should be preserved from template.
        ws_ifs = wb[TARGET_SHEET_TICKETS]
        # Keep existing C3/D5 if they are static in template?
        # Or overwrite just to ensure correctness? 
        # User asked for them, so let's ensure they are there.
        ws_ifs['C3'] = "IFS AMS Workload Summary"
        ws_ifs['D5'] = "For NAFTA Marelli USA"
        
        # Calculate Period String
        period_str = f"Period: {start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
        ws_ifs['D7'] = period_str
        print(f"Updated Dashboard Headers: {period_str}")
        
        # --- TICKETS ---
        print("Processing Tickets...")
        # Read both 2025 and 2026 sheets and combine them
        sheets_to_load = list(dict.fromkeys([target_year, "2025", "2026"]))  # Remove duplicates, preserve order
        df_list = []
        for sheet in sheets_to_load:
            try:
                df_temp = pd.read_excel(SOURCE_FILE, sheet_name=sheet)
                df_temp.columns = df_temp.columns.astype(str).str.strip()
                df_list.append(df_temp)
                print(f"  Loaded sheet: {sheet} ({len(df_temp)} rows)")
            except Exception as e:
                print(f"  Sheet '{sheet}' not found or error: {e}")
        
        if not df_list:
            print("ERROR: No data sheets found")
            return
        
        # Combine and remove duplicates
        df_tickets = pd.concat(df_list, ignore_index=True)
        df_tickets = df_tickets.drop_duplicates()
        print(f"Source Columns: {list(df_tickets.columns)}")
        print(f"Total rows after combining: {len(df_tickets)}")
        
        # NORMALIZE DATE
        renamed = False
        for c in df_tickets.columns:
            if c.lower() in ['date', 'created', 'opened', 'start date']:
                df_tickets.rename(columns={c: 'Date Created'}, inplace=True)
                renamed = True; break
        if not renamed:
             # Look for substring
             for c in df_tickets.columns:
                 if 'date' in c.lower():
                     df_tickets.rename(columns={c: 'Date Created'}, inplace=True)
                     renamed = True; break
        # Fallback Index 1
        if 'Date Created' not in df_tickets.columns and len(df_tickets.columns) > 1:
            print("Fallback: Using Column Index 1 as Date.")
            df_tickets.rename(columns={df_tickets.columns[1]: 'Date Created'}, inplace=True)
            
        if 'Date Created' not in df_tickets.columns:
            print("CRITICAL: Date column missing."); return
            
        df_tickets['Date Created'] = pd.to_datetime(df_tickets['Date Created'], errors='coerce')
        
        # STATUS
        status_col = None
        for c in df_tickets.columns:
            if 'status' in c.lower(): status_col = c; break
            
        # FILTER LOGIC (Open Status Only)
        # User Rule: "Only show tickets with Status = 'Open', exclude future tickets."
        
        # 1. Date Check (exclude future tickets)
        mask_date_valid = df_tickets['Date Created'] <= end_date
        
        # 2. Status must be Open
        if status_col:
            status_clean = df_tickets[status_col].astype(str).str.strip().str.lower()
            mask_open = (status_clean == "open")
        else:
             mask_open = pd.Series([False] * len(df_tickets))
        
        # For stats calculation (tickets created in the week)
        mask_week = (df_tickets['Date Created'] >= start_date) & (df_tickets['Date Created'] <= end_date)
                
        # COMBINE: Must be Open AND not future
        filtered_tickets = df_tickets[mask_open & mask_date_valid]
        print(f" - Found {len(filtered_tickets)} tickets (Status=Open, Date<=EndDate).")
        
        # --- CALCULATE STATS FOR DASHBOARD (D7) ---
        # "calculate how many was open and closed during the previous week period"
        # STRICT INTERPRETATION: Created in that week.
        df_created_in_week = df_tickets[mask_week]
        if status_col:
            stats_status = df_created_in_week[status_col].astype(str).str.strip().str.lower()
            count_closed = stats_status[stats_status.str.contains('close')].count()
            count_open = stats_status[stats_status == 'open'].count()
        else:
            count_closed = 0
            count_open = 0
            
        period_str = f"Period: {start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}   {count_closed} closed, {count_open} open"
        
        # --- DASHBOARD HEADERS (Write Late to include Stats) ---
        ws_ifs = wb[TARGET_SHEET_TICKETS]
        ws_ifs['C3'] = "IFS AMS Workload Summary"
        ws_ifs['D5'] = "For NAFTA Marelli USA"
        
        # Fix for duplicated/overlapping text:
        # Template has "Period: " in D7 and Dates in E7.
        # We are writing the FULL string to D7.
        # We must clear E7 to prevent the old "12/08..." from showing up next to our new text.
        ws_ifs['D7'].value = None
        ws_ifs['E7'].value = None 
        
        ws_ifs['D7'] = period_str
        print(f"Updated Dashboard Headers: {period_str}")
        
        # MAP - Build column mapping from template headers
        def get_col(df, candidates):
            for c in candidates: 
                if c in df.columns: return c
            return None
        
        # Read template headers to build automatic mapping
        ws_tickets = wb[TARGET_SHEET_TICKETS]
        template_headers = {}
        for col_idx in range(1, 27):  # Check columns A-Z
            header_cell = ws_tickets.cell(row=2, column=col_idx)  # Row 2 is header row
            if header_cell.value:
                header_name = str(header_cell.value).strip()
                template_headers[col_idx] = header_name
        
        # Build mapping: explicit mappings first, then auto-map same-named fields
        map_tickets = {}
        
        # Explicit mappings (with fallbacks) - these are source column names
        explicit_source_mappings = {
            'REQ No.': get_col(df_tickets, ['REQ No.', 'REQ No', 'Req No.']),
            'Contact': get_col(df_tickets, ['Contact', 'Requester']),
            'Team member': get_col(df_tickets, ['Team member', 'PIC', 'Team Member']),
            'Date Created': 'Date Created',
            'Remarks': get_col(df_tickets, ['Remarks'])
        }
        
        # Map explicit source columns to template columns by matching header names
        for template_col_idx, template_header in template_headers.items():
            template_header_clean = template_header.strip()
            
            # Handle special mappings
            if template_header_clean == 'ServiceNow Ticket #' or template_header_clean == 'ServiceNow Ticket':
                # Map Ticket No. to ServiceNow Ticket #
                source_col = get_col(df_tickets, ['Ticket No.', 'Ticket No', 'Ticket'])
                if source_col:
                    map_tickets[template_col_idx] = source_col
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source '{source_col}'")
            elif template_header_clean == 'Description':
                if 'Request Detail' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Request Detail'
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source 'Request Detail'")
            elif template_header_clean == 'Type':
                if 'Type' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Type'
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source 'Type'")
            elif template_header_clean == 'PIC':
                if 'Assign To' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Assign To'
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source 'Assign To'")
            elif template_header_clean == 'Received':
                if 'Time - Arrive' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Time - Arrive'
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source 'Time - Arrive'")
            elif template_header_clean == 'Resolved':
                if 'Time - Close' in df_tickets.columns:
                    map_tickets[template_col_idx] = 'Time - Close'
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source 'Time - Close'")
            elif template_header_clean in explicit_source_mappings:
                source_col = explicit_source_mappings[template_header_clean]
                if source_col:
                    map_tickets[template_col_idx] = source_col
                    print(f"Mapped: Column {template_col_idx} ('{template_header}') <- Source '{source_col}'")
        
        # Auto-map: if template header matches source column name exactly (and not already mapped)
        for template_col_idx, template_header in template_headers.items():
            if template_col_idx not in map_tickets:  # Don't override explicit mappings
                template_header_clean = template_header.strip()
                # Check if source has a column with the same name
                if template_header_clean in df_tickets.columns:
                    map_tickets[template_col_idx] = template_header_clean
                    print(f"Auto-mapped: Column {template_col_idx} ('{template_header}') <- Source '{template_header_clean}'")
        
        print(f"Column mappings: {[(k, v) for k, v in sorted(map_tickets.items())]}")
        
        row_idx = DATA_START_ROW
        for _, row in filtered_tickets.iterrows():
            for col_idx in range(8, 27):  # Start from column H (8) to Z (26)
                cell = ws_tickets.cell(row=row_idx, column=col_idx)
                apply_style(cell, ifs_styles.get(col_idx))
                if col_idx in map_tickets and map_tickets[col_idx]:
                    val = row.get(map_tickets[col_idx])
                    if val is not None: cell.value = val
            row_idx += 1
            
        row_idx_tickets = row_idx # Save for table resize
            
        # --- USERS ---
        print("Processing New Users...")
        df_users = pd.read_excel(SOURCE_FILE, sheet_name=SOURCE_SHEET_USERS)
        df_users.columns = df_users.columns.astype(str).str.strip()
        
        renamed_u = False
        for c in df_users.columns:
            if 'date' in c.lower():
                df_users.rename(columns={c: 'Date Created'}, inplace=True)
                renamed_u = True; break
        if not renamed_u and len(df_users.columns) > 1:
            df_users.rename(columns={df_users.columns[1]: 'Date Created'}, inplace=True)
            
        df_users['Date Created'] = pd.to_datetime(df_users['Date Created'], errors='coerce')
        
        mask_date_u = (df_users['Date Created'] >= start_date) & (df_users['Date Created'] <= end_date)
        
        cat_col = None
        for c in df_users.columns:
            if 'category' in c.lower(): cat_col = c; break
            
        if cat_col:
            mask_cat = df_users[cat_col].astype(str).str.strip().str.lower() == "create account"
        else:
            try: mask_cat = df_users.iloc[:, 14].astype(str).str.strip().str.lower() == "create account"
            except: mask_cat = pd.Series([True] * len(df_users))
            
        filtered_users = df_users[mask_date_u & mask_cat]
        print(f" - Found {len(filtered_users)} users.")
        
        ws_users = wb[TARGET_SHEET_USERS]
        row_idx = DATA_START_ROW
        for _, row in filtered_users.iterrows():
            for col_idx in range(1, 20):
                cell = ws_users.cell(row=row_idx, column=col_idx)
                apply_style(cell, user_styles.get(col_idx))
            for i, val in enumerate(row):
                cell = ws_users.cell(row=row_idx, column=i+1)
                cell.value = val
            row_idx += 1
            
        # FINAL STEP: Add Standard AutoFilter (Since we removed the Tables)
        # We need to apply filter to the Header Row + Data range.
        # Header is Row 2.
        
        last_row_map = {
            TARGET_SHEET_TICKETS: row_idx_tickets - 1, 
            TARGET_SHEET_USERS: row_idx # This is the users one
        }

        for ws_name, last_data_row in last_row_map.items():
            if ws_name not in wb.sheetnames: continue
            ws = wb[ws_name]
            
            # Remove any lingering Table objects to prevent corruption errors
            # (We relying on cell styles now)
            if hasattr(ws, 'tables') and ws.tables:
                for tbl_name in list(ws.tables.keys()):
                    del ws.tables[tbl_name]
            
            # Add Standard AutoFilter
            # Range: "A2:Z<last_row>" (Safety max column Z)
            # Find exact max column logic later if needed, but 'P' (16) is max for IFS.
            fill_max_row = max(last_data_row, 2) # At least headers
            ws.auto_filter.ref = f"A2:P{fill_max_row}"
            print(f"Applied AutoFilter to {ws_name}: A2:P{fill_max_row}")
            
        while True:
            try: wb.save(output_filename); print("SUCCESS"); break
            except PermissionError: 
                print("File Open. Close and Press Enter.")
                input()
                
    except Exception as e:
        print(f"CRITICAL ERROR: {e}")
        import traceback; traceback.print_exc()

if __name__ == "__main__":
    generate_report()
